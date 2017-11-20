
from openpyxl import Workbook
from openpyxl import load_workbook

wb = Workbook()
# sheet = wb.active
# sheet['A1'] = '编号'
# sheet['B1'] = '题目'
# sheet['C1'] = '来源'
# sheet['D1'] = '日期'
# sheet['E1'] = '内容'
# sheet['F1'] = 'flag'
# r = range(2,5)
# id = 0
# for i in r:
#     id+=1
#     sheet['A%d'%i] = id
#     sheet['B%d'%i] = artical[0]
#     sheet['C%d'%i] = artical[1]
#     sheet['D%d'%i] = artical[2]
#     sheet['E%d'%i] = content
#     sheet['F%d'%i] = '0'
#
# wb.save('/home/wpf/Desktop/抓取结果.xlsx')

wb1 = load_workbook('/home/wpf/Desktop/抓取结果.xlsx')
sheet = wb1.active
print (wb1.get_sheet_names())
r = range(5,8)
id = 0
for i in r:
    id+=1
    sheet['A%d'%i] = id
    sheet['B%d'%i] = artical[0]
    sheet['C%d'%i] = artical[1]
    sheet['D%d'%i] = artical[2]
    sheet['E%d'%i] = content
    sheet['F%d'%i] = '1'
wb1.save('/home/wpf/Desktop/抓取结果.xlsx')